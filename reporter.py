from openpyxl import Workbook
from openpyxl import load_workbook
from tqdm import tqdm
from dateutil import parser
from openpyxl.styles import Font
from openpyxl.writer.write_only import WriteOnlyCell
import os
import re
from logger import Logger


def get_data(filename):

    wb = load_workbook(filename, read_only=True)
    ws = wb.active
    logs_list = {}
    if ws.cell(row=5, column=1).value == 'Сотрудник'\
            and ws.cell(row=5, column=3).value == 'Таб. номер'\
            and ws.cell(row=5, column=6).value == 'Время события'\
            and ws.cell(row=5, column=7).value == 'Событие'\
            and ws.cell(row=5, column=8).value == 'Точка доступа':
        log_date = ws.cell(row=6, column=6).value.split(' ')[0].strip()
        pbar = tqdm(total=ws.max_row)
        pbar.update(6)
        i = 6
        maximum = ws.max_row
        while i <= maximum:

            tab_number = str(ws.cell(row=i, column=3).value)
            event = ws.cell(row=i, column=7).value.strip()
            log_time = ws.cell(row=i, column=6).value.split(' ')[1].strip()
            ap = ws.cell(row=i, column=8).value.strip()
            employee = ws.cell(row=i, column=1).value
            log_item = {
                    'employee': employee,
                    'log_date': log_date,
                    'tab_number': tab_number,
                    'event': event,
                    'ap': ap,
                    'log_time': log_time
                }

            log = []
            if event == 'Вход':
                log.insert(0, log_item)
            else:
                log.insert(1, log_item)

            if not tab_number in logs_list.keys():
                logs_list[tab_number] = log
            else:
                if event == 'Вход':
                    try:
                        if (parser.parse(log_time) < parser.parse(logs_list[tab_number][0]['log_time'])):
                            logs_list[tab_number].insert(0, log)
                    except IndexError:
                        logs_list[tab_number].insert(0, log)
                elif event == 'Выход':
                    try:
                        if (parser.parse(log_time) > parser.parse(logs_list[tab_number][1]['log_time'])):
                            logs_list[tab_number].insert(1, log[0])
                    except IndexError:
                        logs_list[tab_number].insert(1, log[0])

            pbar.update(1)
            i += 1
    else:
        pass

    return logs_list


def make_report(logs_list, filename):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet('Экспорт')
    font = Font(name='Times New Roman', size=12, bold=True)
    cell1 = WriteOnlyCell(ws, value='Табельный номер')
    cell1.font = font
    ws.column_dimensions['A'].width = 10
    cell2 = WriteOnlyCell(ws, value='ФИО')
    cell2.font = font
    ws.column_dimensions['B'].width = 50
    cell3 = WriteOnlyCell(ws, value='Дата')
    cell3.font = font
    ws.column_dimensions['C'].width = 12
    cell4 = WriteOnlyCell(ws, value='Время первого входа в здание')
    cell4.font = font
    ws.column_dimensions['D'].width = 20
    cell5 = WriteOnlyCell(ws, value='Время последнего выхода из здания')
    cell5.font = font
    ws.column_dimensions['E'].width = 20
    cell6 = WriteOnlyCell(ws, value='Точка входа')
    cell6.font = font
    ws.column_dimensions['F'].width = 15
    cell7 = WriteOnlyCell(ws, value='Точка выхода')
    cell7.font = font
    ws.column_dimensions['G'].width = 15
    ws.append([cell1, cell2, cell3, cell4, cell5, cell6, cell7])

    for log in logs_list.values():
        ws.append([log[0]['tab_number'], log[0]['employee'], log[0]['log_date'], log[0]['log_time'], log[1]['log_time'], log[0]['ap'], log[1]['ap']])

    wb.save('export_' + filename)


if __name__ == '__main__':
    files = []
    BASEDIR = os.path.dirname(os.path.abspath(__file__))
    for file in os.listdir(BASEDIR):
        if re.match('^\d+', file):
            files.append(file)
    if files.__len__() > 0:
        for filename in files:
            if not os.path.isfile(os.path.join(BASEDIR, 'export_' + filename)):
                unsorted_logs = get_data(filename)
                if unsorted_logs.__len__() > 0:
                    make_report(unsorted_logs, filename)
                else:
                    Logger.logger.warning('Файл \"' + filename + '\" не содержит отчета, либо имеет не правильную структуру')
            else:
                Logger.logger.warning('Файл \"' + filename + '\" уже исправлен')
    else:
        Logger.logger.warning('Не найдены файлы отчетов. Убедитесь что названия файлов начинаются с цифры')